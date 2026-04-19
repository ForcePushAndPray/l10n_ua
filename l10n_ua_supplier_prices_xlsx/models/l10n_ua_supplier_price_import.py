import base64
import io
import logging
from datetime import date, datetime, timedelta

import openpyxl

from odoo import models, _
from odoo.exceptions import UserError

from odoo.addons.l10n_ua_supplier_prices_base.tools.transforms import (
    apply_transform, TransformError,
)

_logger = logging.getLogger(__name__)


def _column_letter_to_index(letter: str) -> int:
    """Excel колонка ('A', 'B', 'AC') → 0-based index."""
    letter = letter.upper().strip()
    result = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Invalid column letter: {letter!r}")
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1


class L10nUaSupplierPriceImport(models.Model):
    _inherit = 'l10n_ua.supplier.price.import'

    def _do_parse(self):
        self.ensure_one()
        if self.source_id.parser_type == 'xlsx':
            return self._parse_xlsx()
        return super()._do_parse()

    def _parse_xlsx(self):
        self.ensure_one()
        config = self._get_parser_config()
        sheet_name = config.get('sheet_name')
        header_row = int(config.get('header_row', 1))
        has_header = bool(config.get('has_header', True))
        skip_rows = int(config.get('skip_rows', 0))

        raw_bytes = base64.b64decode(self.raw_file)
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        except Exception as e:
            raise UserError(_("Cannot open XLSX file: %s") % e)

        worksheet = self._select_worksheet(workbook, sheet_name)
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise UserError(_("Worksheet '%s' is empty.") % worksheet.title)

        header_map = {}
        if has_header:
            if header_row < 1 or header_row > len(rows):
                raise UserError(_(
                    "header_row=%(h)s out of range (worksheet has %(n)s rows)",
                    h=header_row, n=len(rows),
                ))
            header_cells = rows[header_row - 1]
            header_map = {
                str(cell).strip(): idx
                for idx, cell in enumerate(header_cells)
                if cell is not None and str(cell).strip()
            }
            data_start = header_row + skip_rows
        else:
            data_start = skip_rows

        mapping = self.source_id.mapping_id
        if not mapping:
            raise UserError(_("Source has no mapping configured."))
        if not mapping.field_ids:
            raise UserError(_("Mapping '%s' has no fields.") % mapping.name)

        deadline = datetime.now() + timedelta(seconds=self.source_id.parse_timeout)
        Line = self.env['l10n_ua.supplier.price.import.line']
        Currency = self.env['res.currency']
        currency_cache = {}
        count = 0

        for row_num, row in enumerate(rows[data_start:], start=data_start + 1):
            if datetime.now() > deadline:
                raise UserError(_(
                    "Parse timeout (%(t)ss) exceeded after %(n)s records.",
                    t=self.source_id.parse_timeout, n=count,
                ))
            if not any(cell is not None for cell in row):
                continue
            vals = {'import_id': self.id, 'sequence': row_num * 10}
            for field in mapping.field_ids:
                raw_value = self._resolve_xlsx_field(row, field.source_path,
                                                    has_header, header_map)
                if raw_value is None or raw_value == '':
                    if field.required and not field.default_value:
                        raise UserError(_(
                            "Required field '%(t)s' missing at row %(r)s (path: %(p)s)",
                            t=field.target, r=row_num, p=field.source_path,
                        ))
                    raw_value = field.default_value
                if raw_value is None or raw_value == '':
                    continue
                try:
                    value = apply_transform(raw_value, field.transform, field.transform_param)
                except TransformError as e:
                    raise UserError(_(
                        "Transform error on field '%(t)s' (row %(r)s): %(e)s",
                        t=field.target, r=row_num, e=str(e),
                    ))
                self._assign_field_value(vals, field.target, value, currency_cache, Currency)

            if not vals.get('supplier_sku'):
                continue
            Line.create(vals)
            count += 1

    def _select_worksheet(self, workbook, sheet_name):
        """Вибрати worksheet за іменем, індексом, або перший доступний."""
        if sheet_name is None or sheet_name == '':
            return workbook.worksheets[0]
        if isinstance(sheet_name, int) or (isinstance(sheet_name, str) and sheet_name.isdigit()):
            idx = int(sheet_name)
            if 0 <= idx < len(workbook.worksheets):
                return workbook.worksheets[idx]
            raise UserError(_("Sheet index %s out of range") % idx)
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        raise UserError(_(
            "Sheet '%(s)s' not found. Available: %(a)s",
            s=sheet_name, a=', '.join(workbook.sheetnames),
        ))

    def _resolve_xlsx_field(self, row, source_path, has_header, header_map):
        """Отримати значення з row tuple за source_path.

        has_header=True: source_path = назва колонки, шукаємо в header_map
        has_header=False: source_path = літера ('A') або індекс ('0')
        """
        if not source_path:
            return None
        if has_header:
            idx = header_map.get(source_path.strip())
            if idx is None:
                return None
        else:
            try:
                idx = int(source_path)
            except (ValueError, TypeError):
                try:
                    idx = _column_letter_to_index(source_path)
                except ValueError:
                    return None
        if not (0 <= idx < len(row)):
            return None
        cell = row[idx]
        if isinstance(cell, datetime):
            return cell.date()
        return cell
