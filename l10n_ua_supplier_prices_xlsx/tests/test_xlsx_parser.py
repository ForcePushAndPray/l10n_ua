"""Тести XLSX парсера. Файли генеруються openpyxl in-memory."""

import base64
import io
import json
from datetime import date

import openpyxl

from odoo.tests import TransactionCase, tagged

from odoo.addons.l10n_ua_supplier_prices_xlsx.models.l10n_ua_supplier_price_import import (
    _column_letter_to_index,
)


def make_xlsx(rows, sheet_name='Prices'):
    """Створити XLSX bytes з 2D-списку рядків."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@tagged('post_install', '-at_install')
class TestXlsxParser(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.partner = cls.env['res.partner'].create({
            'name': 'XLSX Supplier',
            'supplier_rank': 1,
        })
        cls.uah = cls.env.ref('base.UAH')

    def _make_source(self, parser_config, with_header=True, source_paths=None):
        if source_paths is None:
            source_paths = {
                'sku': 'SKU' if with_header else 'A',
                'name': 'Name' if with_header else 'B',
                'price': 'Price' if with_header else 'C',
            }
        mapping = self.env['l10n_ua.supplier.price.mapping'].create({
            'name': 'XLSX Test',
            'parser_config': json.dumps(parser_config),
            'field_ids': [
                (0, 0, {'target': 'sku', 'source_path': source_paths['sku'],
                        'required': True}),
                (0, 0, {'target': 'name', 'source_path': source_paths['name']}),
                (0, 0, {'target': 'price', 'source_path': source_paths['price'],
                        'required': True, 'transform': 'to_float'}),
            ],
        })
        return self.env['l10n_ua.supplier.price.source'].create({
            'name': 'XLSX Source',
            'partner_id': self.partner.id,
            'parser_type': 'xlsx',
            'fetcher_type': 'manual',
            'mapping_id': mapping.id,
            'currency_default_id': self.uah.id,
        })

    def _make_import(self, source, content_bytes, filename='test.xlsx'):
        imp = self.env['l10n_ua.supplier.price.import'].create({
            'source_id': source.id,
            'raw_file': base64.b64encode(content_bytes),
            'raw_filename': filename,
        })
        imp._set_state('fetching')
        imp._set_state('fetched')
        return imp

    # === unit test для хелпера ===

    def test_column_letter_to_index(self):
        self.assertEqual(_column_letter_to_index('A'), 0)
        self.assertEqual(_column_letter_to_index('B'), 1)
        self.assertEqual(_column_letter_to_index('Z'), 25)
        self.assertEqual(_column_letter_to_index('AA'), 26)
        self.assertEqual(_column_letter_to_index('AB'), 27)
        self.assertEqual(_column_letter_to_index('BA'), 52)

    def test_column_letter_invalid(self):
        with self.assertRaises(ValueError):
            _column_letter_to_index('A1')

    # === integration ===

    def test_xlsx_with_header(self):
        source = self._make_source({'has_header': True})
        content = make_xlsx([
            ['SKU', 'Name', 'Price'],
            ['SKU001', 'Виріб А', 100.50],
            ['SKU002', 'Виріб Б', 250.00],
        ])
        imp = self._make_import(source, content)
        imp.action_parse()
        self.assertEqual(imp.state, 'parsed')
        self.assertEqual(len(imp.line_ids), 2)
        self.assertEqual(imp.line_ids[0].supplier_sku, 'SKU001')
        self.assertEqual(imp.line_ids[0].supplier_name, 'Виріб А')
        self.assertEqual(imp.line_ids[0].price, 100.50)

    def test_xlsx_without_header_letter_path(self):
        source = self._make_source(
            {'has_header': False},
            with_header=False,
            source_paths={'sku': 'A', 'name': 'B', 'price': 'C'},
        )
        content = make_xlsx([
            ['SKU001', 'Product A', 100.50],
            ['SKU002', 'Product B', 250.00],
        ])
        imp = self._make_import(source, content)
        imp.action_parse()
        self.assertEqual(len(imp.line_ids), 2)
        self.assertEqual(imp.line_ids[0].supplier_sku, 'SKU001')

    def test_xlsx_without_header_index_path(self):
        source = self._make_source(
            {'has_header': False},
            with_header=False,
            source_paths={'sku': '0', 'name': '1', 'price': '2'},
        )
        content = make_xlsx([
            ['SKU001', 'Product A', 100.50],
        ])
        imp = self._make_import(source, content)
        imp.action_parse()
        self.assertEqual(imp.line_ids.supplier_sku, 'SKU001')

    def test_xlsx_specific_sheet(self):
        wb = openpyxl.Workbook()
        wb.active.title = 'Empty'
        ws = wb.create_sheet('PriceList')
        ws.append(['SKU', 'Name', 'Price'])
        ws.append(['SKU001', 'A', 10.0])
        buf = io.BytesIO()
        wb.save(buf)
        source = self._make_source({'has_header': True, 'sheet_name': 'PriceList'})
        imp = self._make_import(source, buf.getvalue())
        imp.action_parse()
        self.assertEqual(len(imp.line_ids), 1)

    def test_xlsx_skip_rows_metadata(self):
        source = self._make_source({'has_header': True, 'skip_rows': 2})
        content = make_xlsx([
            ['SKU', 'Name', 'Price'],
            ['Skip me 1', 'Skip me 1', 'Skip me 1'],
            ['Skip me 2', 'Skip me 2', 'Skip me 2'],
            ['SKU001', 'Real', 100.0],
        ])
        imp = self._make_import(source, content)
        imp.action_parse()
        self.assertEqual(len(imp.line_ids), 1)
        self.assertEqual(imp.line_ids.supplier_sku, 'SKU001')

    def test_xlsx_header_row_2(self):
        """Перший рядок — meta, header у рядку 2."""
        source = self._make_source({'has_header': True, 'header_row': 2})
        content = make_xlsx([
            ['Прайс на 2026-04-19', '', ''],
            ['SKU', 'Name', 'Price'],
            ['SKU001', 'Real', 100.0],
        ])
        imp = self._make_import(source, content)
        imp.action_parse()
        self.assertEqual(len(imp.line_ids), 1)

    def test_xlsx_skip_empty_rows(self):
        source = self._make_source({'has_header': True})
        content = make_xlsx([
            ['SKU', 'Name', 'Price'],
            ['SKU001', 'A', 10.0],
            [None, None, None],
            ['SKU002', 'B', 20.0],
        ])
        imp = self._make_import(source, content)
        imp.action_parse()
        self.assertEqual(len(imp.line_ids), 2)

    def test_xlsx_numeric_price_no_transform(self):
        """Excel вже зберігає 100.50 як float — to_float не ламається."""
        source = self._make_source({'has_header': True})
        content = make_xlsx([
            ['SKU', 'Name', 'Price'],
            ['X', 'Y', 99.99],
        ])
        imp = self._make_import(source, content)
        imp.action_parse()
        self.assertEqual(imp.line_ids.price, 99.99)

    def test_xlsx_sheet_not_found(self):
        source = self._make_source({'has_header': True, 'sheet_name': 'Nonexistent'})
        content = make_xlsx([['SKU', 'Name', 'Price'], ['X', 'Y', 1]])
        imp = self._make_import(source, content)
        imp.action_parse()
        self.assertEqual(imp.state, 'error')
        self.assertIn('not found', imp.error_log)

    def test_xlsx_invalid_file(self):
        source = self._make_source({'has_header': True})
        imp = self._make_import(source, b'not an xlsx file')
        imp.action_parse()
        self.assertEqual(imp.state, 'error')
        self.assertIn('Cannot open XLSX', imp.error_log)

    def test_xlsx_missing_required_field(self):
        source = self._make_source({'has_header': True})
        content = make_xlsx([
            ['SKU', 'Name', 'Price'],
            [None, 'Has name', 100.0],  # SKU missing
        ])
        imp = self._make_import(source, content)
        imp.action_parse()
        self.assertEqual(imp.state, 'error')
        self.assertIn('missing', imp.error_log)
